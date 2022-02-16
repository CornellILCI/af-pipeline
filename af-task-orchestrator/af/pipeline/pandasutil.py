import os
from typing import Iterable

import openpyxl
import pandas as pd
from af.pipeline import exceptions
from openpyxl.utils import get_column_letter


def df_keep_columns(df: pd.DataFrame, columns_to_keep: Iterable[str]) -> pd.DataFrame:
    """Keeps only columns in given set.

    Keeps only columns in input set and drops columns not in the given set.
    Unmapped column in input set is ignored.

    A simple df[keep_columns] will fail when one of the keep column missing in df.

    Args:
        df:
            Input dataframe
        columns_to_keep:
            set of columns to keep in dataframe.

    Returns:
        df with only columns to keep.
    """

    # convert to set
    columns_to_keep_set = set(columns_to_keep)

    columns_to_drop = set(df.columns) - columns_to_keep_set
    columns_to_keep_set = set(df.columns) - columns_to_drop

    # to make sure order
    eligible_columns = []
    for column in columns_to_keep:
        if column in columns_to_keep_set:
            eligible_columns.append(column)

    df = df[eligible_columns]

    return df


def set_columns_as_numeric(df: pd.DataFrame, columns: list):

    for column in columns:
        if column in df:
            df[column] = pd.to_numeric(df[column], errors="ignore")

    return df


def append_df_to_excel(filename, df, sheet_name="Sheet1", startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Args:
        filename: File path or existing ExcelWriter
                  (Example: '/path/to/file.xlsx')
        df: DataFrame to save to workbook
        sheet_name: Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow: upper left cell row to dump data frame.
                  Per default (startrow=None) calculate the last row
                  in the existing DF and write to the next row...
        truncate_sheet: truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                         [can be a dictionary]

    (ref) [MaxU](https://stackoverflow.com/questions/38074678/
    append-existing-excel-sheet-with-new-dataframe-using-python-pandas)
    """

    # by default ignore df indices
    if "index" not in to_excel_kwargs:
        to_excel_kwargs["index"] = False

    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename, sheet_name=sheet_name, startrow=startrow if startrow is not None else 0, **to_excel_kwargs
        )

        # set optimal width
        xl_book = openpyxl.load_workbook(filename)
        set_optimal_columns_widths(xl_book, sheet_name, df)
        xl_book.save(filename)

        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl", mode="a")

    # try to open an existing workbook
    writer.book = openpyxl.load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    if sheet_name in writer.sheets and startrow > 1 and "header" not in to_excel_kwargs:
        to_excel_kwargs["header"] = False

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    set_optimal_columns_widths(writer.book, sheet_name, df)

    # save the workbook
    writer.save()


def set_optimal_columns_widths(xl_book, sheet_name, df):

    ws = xl_book[sheet_name]
    for i, column_name in enumerate(df.columns):

        # add 5 units to give more space
        optimal_width = max(len(column_name), df[column_name].map(str).map(len).max()) + 5

        # openpyxl column index is 1 based
        column_letter = get_column_letter(i + 1)

        ws.column_dimensions[column_letter].width = optimal_width
